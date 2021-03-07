import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Get html text by using requests library
req = requests.get("http://leonkong.surge.sh")
html = req.text

# Parse html text by using Beautiful Soup 4
soup = BeautifulSoup(html, 'html.parser')
post_titles = soup.select('a > div > h3')

# Delete date info from the title and add it to results
titles = []
dates = []
for title in post_titles:
    raw_title = title.text.split('-')
    titles.append(raw_title[0])
    dates.append(raw_title[1])

# Save data in a .xlsx excel format
wb = Workbook()
ws = wb.active # Grab the active worksheet

# Write data on worksheet
ws['A1'] = 'Title'
ws['B1'] = 'Date'


for i in range(len(titles)):
    pos = str(i+2)
    ws['A' + pos] = titles[i]
    ws['B' + pos] = dates[i]

# Save the file
wb.save("posts.xlsx")