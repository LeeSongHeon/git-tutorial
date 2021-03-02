import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


req = requests.get("http://leonkong.surge.sh")
html = req.text


soup = BeautifulSoup(html, 'html.parser')
post_titles = soup.select('a > div > h3')


titles = []
dates = []
for title in post_titles:
    raw_title = title.text.split('-')
    titles.append(raw_title[0])
    dates.append(raw_title[1])


wb = Workbook()
ws = wb.active 


ws['A1'] = 'Title'
ws['B1'] = 'Date'


for i in range(len(titles)):
    pos = str(i+2)
    ws['A' + pos] = titles[i]
    ws['B' + pos] = dates[i]


wb.save("posts.xlsx")